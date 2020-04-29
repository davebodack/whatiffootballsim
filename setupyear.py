from openpyxl import load_workbook
import sys

filepath = "./All-Timers League.xlsx"
wb = load_workbook(filepath)

#Set up teams in each division
divsheet = wb["Divisions"]
lastyear = int(sys.argv[1]) - 1
pointdiffs = wb["Year " + str(lastyear) + " Point Differentials"]
for i in range(1, 33):
	divsheet.cell(row = i, column = 2).value = pointdiffs.cell(row = i+1, column = 2).value 


#Set up schedule sheet (I know this way is stupid but so is openpyxl)
divcellrowcords = [2,8,8,2,2,8,8,2,3,9,9,3,3,9,9,3,4,10,10,4,4,10,10,4,5,11,11,5,5,11,11,5]
divcellcolcords = [4,4,7,7,6,6,5,5,5,5,6,6,7,7,4,4,4,4,7,7,6,6,5,5,5,5,6,6,7,7,4,4]
teamorder = [""] * 32
for i in range(8):
	for j in range(4):
		divsheet.cell(row = divcellrowcords[i*4+j], column = divcellcolcords[i*4+j]).value = divsheet.cell(row = i*4+j+1, column = 2).value
		
for i in range(8):
	for j in range(4):
		if i < 4:
			teamorder[i*4+j] = divsheet.cell(row = j+2, column = i+4).value
		else:
			teamorder[i*4+j] = divsheet.cell(row = j+8, column = i).value
			
schedulename = "Year " + sys.argv[1] + " Schedule"
wb.create_sheet(schedulename)
schedule = wb[schedulename]
for i in range(2,66,2):
	#schedule.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+1)
	schedule.cell(row = 1, column = i).value = teamorder[int((i-2)/2)]


#Create standings sheet
divsheetname = "Year " + sys.argv[1] + " Division Standings"
wb.create_sheet(divsheetname)
standingssheet = wb[divsheetname]


#Set up standings template
divcounter = 1
for i in range(1,11,3):
	#schedule.merge_cells(start_row = 1, start_column = i, end_row = 1, end_column = i+1)
	standingssheet.cell(row = 1, column = i).value = "Division " + str(divcounter)
	divcounter += 1

divcounter = 5
for i in range(1,11,3):
	#schedule.merge_cells(start_row = 7, start_column = i, end_row = 7, end_column = i+1)
	standingssheet.cell(row = 7, column = i).value = "Division " + str(divcounter)
	divcounter += 1


#For now I have to manually create the empty bracket with the right sheet name but I'll figure that out eventually


#Set up point differentials sheet
diffname = "Year " + sys.argv[1] + " Point Differentials"
wb.create_sheet(diffname)
diffsheet = wb[diffname]

for i in range(1, 33):
	diffsheet.cell(i+1,1).value = str(i) + "."
diffsheet.cell(1,3).value = "Points Gained"
diffsheet.cell(1,4).value = "Points Allowed"
diffsheet.cell(1,5).value = "Point Differential"
diffsheet.cell(1,6).value = "Record"


wb.save(filepath)