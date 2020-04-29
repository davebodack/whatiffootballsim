from openpyxl import load_workbook
from datetime import datetime
import requests
import random
import sys


class Team(object):

    def __init__(self):
        self.name = ""
        self.wins = 0
        self.losses = 0
        self.ties = 0
        self.record = ""
        self.pointsscored = 0
        self.pointsallowed = 0
        self.playoffseed = 0
        self.opponents = ["" for x in range(16)]
        self.homefield = ["" for x in range(16)]
        self.results = ["" for x in range(16)]

    def __eq__(self, obj):
        return isinstance(obj, Team) and obj.name == self.name
        

def index_2d(list, name):
    for i, j in enumerate(list):
        for x in j:
            if x.name == name:
                return (i, j.index(x))


def winsort(list):
    temp = Team()
    for i in range(1, len(list)-1):
        for j in range(0, (len(list) - (i + 1))):
            if (list[i].wins < list[i+1].wins) or ((list[i].wins == list[i+1].wins) and (list[i].ties < list[i+1].ties)):
                temp = list[i]
                list[i] = list[i+1]
                list[i+1] = temp


def team_tiebreaker(team1, team2):

    #First look at two teams' record against each other
    # print(team1.name + ": " + str(team1.wins) + "-" + str(team1.losses))
    # print(team2.name + ": " + str(team2.wins) + "-" + str(team2.losses))
    team1netrecord = 0
    team2netrecord = 0
    for i in range(16):
        if team1.opponents[i] == team2.name:
            if team1.results[i] == "W":
                team1netrecord += 1
                team2netrecord -= 1
            elif team1.results[i] == "L":
                team1netrecord -= 1
                team2netrecord += 1

    if team1netrecord > team2netrecord:
        # print("Broke on record against each other")
        return team1.name
    elif team1netrecord < team2netrecord:
        # print("Broke on record against each other")
        return team2.name
    else:

        #Then look at two teams' records against other teams in their division
        team1netrecord = 0
        team2netrecord = 0
        for i in range(3):
            if team1.results[i] == "W":
                team1netrecord += 1
            elif team1.results[i] == "L":
                team1netrecord -= 1
            if team2.results[i] == "W":
                team2netrecord += 1
            elif team2.results[i] == "L":
                team2netrecord -= 1

        for i in range(13,16):
            if team1.results[i] == "W":
                team1netrecord += 1
            elif team1.results[i] == "L":
                team1netrecord -= 1
            if team2.results[i] == "W":
                team2netrecord += 1
            elif team2.results[i] == "L":
                team2netrecord -= 1

        if team1netrecord > team2netrecord:
            # print("Broke on division record")
            return team1.name
        elif team1netrecord < team2netrecord:
            # print("Broke on division record")
            return team2.name
        else:

            #Then check each team's record against their common opponents
            team1netrecord = 0
            team2netrecord = 0
            team1set = set(team1.opponents)
            team2set = set(team2.opponents)
            sameops = team1set.intersection(team2set)
            for i in sameops:
                if team1.results[team1.opponents.index(i)] == "W":
                    team1netrecord += 1
                elif team1.results[team1.opponents.index(i)] == "L":
                    team1netrecord -= 1

                if team2.results[team2.opponents.index(i)] == "W":
                    team2netrecord += 1
                elif team2.results[team2.opponents.index(i)] == "L":
                    team2netrecord -= 1

            if team1netrecord > team2netrecord:
                # print("Swapped on record against common opponents")
                return team1.name
            elif team1netrecord < team2netrecord:
                # print("Swapped on record against common opponents")
                return team2.name
            else:

                #Then check each team's record in their conference
                team1netrecord = 0
                team2netrecord = 0
                for i in range(9):

                    if team1.results[i] == "W":
                        team1netrecord += 1
                    elif team1.results[i] == "L":
                        team1netrecord -= 1

                    if team2.results[i] == "W":
                        team2netrecord += 1
                    elif team2.results[i] == "L":
                        team2netrecord -= 1

                for i in range(13,16):

                    if team1.results[i] == "W":
                        team1netrecord += 1
                    elif team1.results[i] == "L":
                        team1netrecord -= 1

                    if team2.results[i] == "W":
                        team2netrecord += 1
                    elif team2.results[i] == "L":
                        team2netrecord -= 1

                if team1netrecord > team2netrecord:
                    # print("Broke on record in conference")
                    return team1.name
                elif team1netrecord < team2netrecord:
                    # print("Broke on record in conference")
                    return team2.name
                else:

                    #Then look at each team's point differential
                    if (team1.pointsscored - team1.pointsallowed) > (team2.pointsscored - team2.pointsallowed):
                        # print("Broke on point differential")
                        return team1.name
                    elif (team2.pointsscored - team2.pointsallowed) > (team1.pointsscored - team1.pointsallowed):
                        # print("Broke on point differential")
                        return team2.name
                    else:

                        #Then pick one of the teams at random
                        # print("Broke randomly")
                        return random.choice([team1.name, team2.name])


#Bubble sort function to sort teams by playoff eligibility using above function team_tiebreaker
#Code taken from stackabuse.com/sorting-algorithms-in-python/
def playoffsort(list):

    length = len(list)
    for i in range(length):
        for j in range(0, length-i-1):
            # input("Press Enter:")
            if ((list[j].wins > list[j+1].wins) or ((list[j].wins == list[j+1].wins) and (list[j].ties > list[j+1].ties))):
                list[j], list[j+1] = list[j+1], list[j]
                # print("Swapped on record")
            elif ((list[j].wins == list[j+1].wins) and (list[j].ties == list[j+1].ties)):
                if team_tiebreaker(list[j], list[j+1]) == list[j].name:
                    list[j], list[j+1] = list[j+1], list[j]
                    # print("...okay we actually swapped")
            # else:
            #     print("We didn't swap")
            # print("List: ")
            # for k in range(length):
            #     print(list[k].name + ": " + list[k].record)
            # print("")

    list.reverse()
    # print("")
    # for i in range(length):
    #     print(list[i].name + ": " + list[i].record)
    # print("Finished sorting the list")
    # print("")
  


def run_game(hometeam, awayteam):

    hometeamfullsplit = hometeam.split(" ")
    awayteamfullsplit = awayteam.split(" ")
    requeststr = "https://www.whatifsports.com/NFL/default.asp?hSeason="
    requeststr += hometeamfullsplit[0]
    requeststr += "&hteam="
    for n in range(1, len(hometeamfullsplit)):
        if n + 1 < len(hometeamfullsplit):
            requeststr += (hometeamfullsplit[n] + "+")
        else:
            requeststr += (hometeamfullsplit[n] + "&")
    requeststr += "vSeason="
    requeststr += awayteamfullsplit[0]
    requeststr += "&vteam="
    for n in range(1, len(awayteamfullsplit)):
        if n + 1 < len(awayteamfullsplit):
            requeststr += (awayteamfullsplit[n] + "+")
        else:
            requeststr += (awayteamfullsplit[n])
    requestget = requests.get(requeststr)
    gameid = requestget.text.split("GameID=")
    gameid = gameid[1].split("&", 1)
    gameid = gameid[0]

    requestboxscore = "https://www.whatifsports.com/NFL/pbp.asp?gameid="
    requestboxscore += gameid
    requestboxscore += "&qtr=4&teamfee=-1"
    boxscore = requests.get(requestboxscore)
    boxscore = boxscore.text
    hometeampartsplit = hometeam.rsplit(" ", 1)
    hometeampartsplit = hometeampartsplit[0] + ":"
    awayteampartsplit = awayteam.rsplit(" ", 1)
    awayteampartsplit = awayteampartsplit[0] + ":"

    if not ("Overtime 1" in boxscore):
        boxscoreawaysplit = boxscore.split(awayteampartsplit)[1]
    else:
        boxscoreawaysplit = boxscore.split(awayteampartsplit)[2]
        print("Overtime")

    awayscore = boxscoreawaysplit.split(" ", 2)[1]
    homescore = boxscoreawaysplit[boxscoreawaysplit.find(hometeampartsplit) + len(hometeampartsplit):boxscoreawaysplit.rfind("</font>")].strip()
    awayscore = int(awayscore)
    homescore = int(homescore)

    return [homescore, awayscore, gameid]


#All right, let's get this show on the road.
filepath = "./All-Timers League.xlsx"
wb = load_workbook(filepath)
sheet = wb["Divisions"]

#Read the team names into the Team objects in the array 'teams'.
teams = [[Team() for i in range(4)] for j in range(8)]

for i in range(8):
    for j in range(4):
        if i < 4:
            teams[i][j].name = sheet.cell(row = j+2, column = i+4).value
        else:
            teams[i][j].name = sheet.cell(row = j+8, column = i).value
        print(teams[i][j].name)
    print("")

schedulesheetname = "Year " + sys.argv[1] + " Schedule"
schedulesheet = wb[schedulesheetname]

indices = [1,2,3]
afcdivisionorder = [1,2,3] 
nfcdivisionorder = [1,2,3]
afcdivnumorders = [0] * 4
nfcdivnumorders = [0] * 4
for i in range(4):
    afcdivnumorders[i] = [0,1,2,3]
    nfcdivnumorders[i] = [0,1,2,3]

for i in range(8):
    
    m1t1home = random.choice([True, False])
    if m1t1home == True:
        m1t2home = False
    else:
        m1t2home = True
    m2t1home = random.choice([True, False])
    if m2t1home == True:
        m2t2home = False
    else:
        m2t2home = True
    
    for j in range(3):
        if j % 2 == 0:
            teams[i][0].homefield[j] = m1t1home
            teams[i][indices[0]].homefield[j] = m1t2home
        else:
            teams[i][0].homefield[j] = not m1t1home
            teams[i][indices[0]].homefield[j] = not m1t2home

    for j in range(13,16):
        if j % 2 == 0:
            teams[i][0].homefield[j] = m1t1home
            teams[i][indices[0]].homefield[j] = m1t2home
        else:
            teams[i][0].homefield[j] = not m1t1home
            teams[i][indices[0]].homefield[j] = not m1t2home

    for j in range(3):
        if j % 2 == 0:
            teams[i][indices[1]].homefield[j] = m2t1home
            teams[i][indices[2]].homefield[j] = m2t2home
        else:
            teams[i][indices[1]].homefield[j] = not m2t1home
            teams[i][indices[2]].homefield[j] = not m2t2home

    for j in range(13,16):
        if j % 2 == 0:
            teams[i][indices[1]].homefield[j] = m2t1home
            teams[i][indices[2]].homefield[j] = m2t2home
        else:
            teams[i][indices[1]].homefield[j] = not m2t1home
            teams[i][indices[2]].homefield[j] = not m2t2home

    #Set teams' divisional opponents for Weeks 1-3
    random.shuffle(indices)
    t1copw2 = random.choice([indices[1], indices[2]])
    if t1copw2 == indices[1]:
        t1notcopw2 = indices[2]
    else:
        t1notcopw2 = indices[1]
    teams[i][0].opponents[0] = teams[i][indices[0]].name
    teams[i][indices[0]].opponents[0] = teams[i][0].name
    teams[i][indices[1]].opponents[0] = teams[i][indices[2]].name
    teams[i][indices[2]].opponents[0] = teams[i][indices[1]].name
    teams[i][0].opponents[1] = teams[i][t1copw2].name
    teams[i][t1copw2].opponents[1] = teams[i][0].name
    teams[i][indices[0]].opponents[1] = teams[i][t1notcopw2].name
    teams[i][t1notcopw2].opponents[1] = teams[i][indices[0]].name
    teams[i][0].opponents[2] = teams[i][t1notcopw2].name
    teams[i][t1notcopw2].opponents[2] = teams[i][0].name
    teams[i][indices[0]].opponents[2] = teams[i][t1copw2].name
    teams[i][t1copw2].opponents[2] = teams[i][indices[0]].name

    #Set teams' divisional opponents for Weeks 14-16
    random.shuffle(indices)
    t1copw2 = random.choice([indices[1], indices[2]])
    if t1copw2 == indices[1]:
        t1notcopw2 = indices[2]
    else:
        t1notcopw2 = indices[1]
    teams[i][0].opponents[13] = teams[i][indices[0]].name
    teams[i][indices[0]].opponents[13] = teams[i][0].name
    teams[i][indices[1]].opponents[13] = teams[i][indices[2]].name
    teams[i][indices[2]].opponents[13] = teams[i][indices[1]].name
    teams[i][0].opponents[14] = teams[i][t1copw2].name
    teams[i][t1copw2].opponents[14] = teams[i][0].name
    teams[i][indices[0]].opponents[14] = teams[i][t1notcopw2].name
    teams[i][t1notcopw2].opponents[14] = teams[i][indices[0]].name
    teams[i][0].opponents[15] = teams[i][t1notcopw2].name
    teams[i][t1notcopw2].opponents[15] = teams[i][0].name
    teams[i][indices[0]].opponents[15] = teams[i][t1copw2].name
    teams[i][t1copw2].opponents[15] = teams[i][indices[0]].name
    
random.shuffle(afcdivisionorder)
random.shuffle(nfcdivisionorder)
for i in range(4):
    random.shuffle(afcdivnumorders[i])
    random.shuffle(nfcdivnumorders[i])

randbool = random.choice([True, False])

for i in range(3,9,2):

    for j in range(4):

        if i == 3:
            teams[0][afcdivnumorders[0][j]].opponents[i] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
            teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i] = teams[0][afcdivnumorders[0][j]].name 
            teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
            teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
            teams[4][nfcdivnumorders[0][j]].opponents[i] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].name
            teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i] = teams[4][nfcdivnumorders[0][j]].name
            teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
            teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[afcdivisionorder[1]][j]].name         
        
            if not(j == 3):
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j+1]].name
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][j+1]].name 
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j+1]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j+1]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j+1]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][j+1]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j+1]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[afcdivisionorder[1]][j+1]].name         
            else:
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][0]].name
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][0]].name 
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][0]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][0]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][0]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][0]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][0]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[afcdivisionorder[1]][0]].name         

            for k in range(2):
                teams[0][afcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].homefield[i+k] = not randbool
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].homefield[i+k] = not randbool
                teams[4][nfcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].homefield[i+k] = not randbool
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].homefield[i+k] = not randbool
                randbool = not randbool

        elif i == 5:
            teams[0][afcdivnumorders[0][j]].opponents[i] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
            teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i] = teams[0][afcdivnumorders[0][j]].name 
            teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
            teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
            teams[4][nfcdivnumorders[0][j]].opponents[i] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name
            teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i] = teams[4][nfcdivnumorders[0][j]].name
            teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
            teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][j]].name         
       
            if not(j == 3):
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j+1]].name
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][j+1]].name 
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j+1]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j+1]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j+1]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][j+1]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j+1]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][j+1]].name         
            else:
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][0]].name
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][0]].name 
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][0]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][0]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][0]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][0]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][0]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][0]].name         

            for k in range(2):
                teams[0][afcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].homefield[i+k] = not randbool
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].homefield[i+k] = not randbool
                teams[4][nfcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].homefield[i+k] = not randbool
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].homefield[i+k] = not randbool
                randbool = not randbool

        else:
            teams[0][afcdivnumorders[0][j]].opponents[i] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
            teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i] = teams[0][afcdivnumorders[0][j]].name 
            teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
            teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
            teams[4][nfcdivnumorders[0][j]].opponents[i] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
            teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i] = teams[4][nfcdivnumorders[0][j]].name
            teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name
            teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][j]].name         
     
            if not(j == 3):
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j+1]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][j+1]].name 
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j+1]].name
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j+1]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j+1]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][j+1]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j+1]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][j+1]].name         
            else:
                teams[0][afcdivnumorders[0][j]].opponents[i+1] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][0]].name
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[i+1] = teams[0][afcdivnumorders[0][0]].name 
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[i+1] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][0]].name
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[i+1] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][0]].name
                teams[4][nfcdivnumorders[0][j]].opponents[i+1] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][0]].name
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[i+1] = teams[4][nfcdivnumorders[0][0]].name
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[i+1] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][0]].name
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[i+1] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[afcdivisionorder[0]][0]].name         
         
            for k in range(2):
                teams[0][afcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].homefield[i+k] = not randbool
                teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].homefield[i+k] = randbool
                teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].homefield[i+k] = not randbool
                teams[4][nfcdivnumorders[0][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].homefield[i+k] = not randbool
                teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].homefield[i+k] = randbool
                teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].homefield[i+k] = not randbool
                randbool = not randbool

        randbool = random.choice([True, False])

for j in range(4):

    #Matchups for Week 10
    teams[0][afcdivnumorders[0][j]].opponents[9] = teams[4][nfcdivnumorders[0][j]].name
    teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[9] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].name 
    teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[9] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name
    teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[9] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
    teams[4][nfcdivnumorders[0][j]].opponents[9] = teams[0][afcdivnumorders[0][j]].name
    teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[9] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
    teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[9] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
    teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[9] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name

    #Matchups for Week 11
    teams[0][afcdivnumorders[0][j]].opponents[10] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].name
    teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[10] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name 
    teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[10] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
    teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[10] = teams[4][nfcdivnumorders[0][j]].name
    teams[4][nfcdivnumorders[0][j]].opponents[10] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
    teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[10] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
    teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[10] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
    teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[10] = teams[0][afcdivnumorders[0][j]].name

    #Matchups for Week 12
    teams[0][afcdivnumorders[0][j]].opponents[11] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name
    teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[11] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name 
    teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[11] = teams[4][nfcdivnumorders[0][j]].name
    teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[11] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].name
    teams[4][nfcdivnumorders[0][j]].opponents[11] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name
    teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[11] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
    teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[11] = teams[0][afcdivnumorders[0][j]].name
    teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[11] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name

    #Matchups for Week 13
    teams[0][afcdivnumorders[0][j]].opponents[12] = teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].name
    teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].opponents[12] = teams[4][nfcdivnumorders[0][j]].name 
    teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].opponents[12] = teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].name
    teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].opponents[12] = teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].name
    teams[4][nfcdivnumorders[0][j]].opponents[12] = teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].name
    teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].opponents[12] = teams[0][afcdivnumorders[0][j]].name
    teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].opponents[12] = teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].name
    teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].opponents[12] = teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].name

    for k in range(9,13):
        teams[0][afcdivnumorders[0][j]].homefield[k] = randbool
        teams[afcdivisionorder[0]][afcdivnumorders[afcdivisionorder[0]][j]].homefield[k] = randbool
        teams[afcdivisionorder[1]][afcdivnumorders[afcdivisionorder[1]][j]].homefield[k] = randbool
        teams[afcdivisionorder[2]][afcdivnumorders[afcdivisionorder[2]][j]].homefield[k] = randbool
        teams[4][nfcdivnumorders[0][j]].homefield[k] = not randbool
        teams[nfcdivisionorder[0]+4][nfcdivnumorders[nfcdivisionorder[0]][j]].homefield[k] = not randbool
        teams[nfcdivisionorder[1]+4][nfcdivnumorders[nfcdivisionorder[1]][j]].homefield[k] = not randbool
        teams[nfcdivisionorder[2]+4][nfcdivnumorders[nfcdivisionorder[2]][j]].homefield[k] = not randbool

        randbool = not randbool


#Writes the opponents of each team to the proper Schedule sheet in the workbook
for div in range(8):
    for team in range(4):
        for i in range(2,18):
            if teams[div][team].homefield[i-2] == True:
                schedulesheet.cell(row = i, column = div*8 + team*2 + 2, value = "vs. " + teams[div][team].opponents[i-2])
            else:
                schedulesheet.cell(row = i, column = div*8 + team*2 + 2, value = "@ " + teams[div][team].opponents[i-2])


#Constructs the proper URLs, sends the requests, and parses them for the game scores
#Calls run_game function to do so
#RENAME HOMETEAM AND AWAYTEAM, THEYRE NOT ALWAYS HOME AND AWAY
for i in range(2,66,2):
    hometeam = schedulesheet.cell(row = 1, column = i).value
    for j in range(2,18):
        if not schedulesheet.cell(row = j, column = i+1).value:
            awayteam = schedulesheet.cell(row = j, column = i).value
            print("*** " + hometeam + " ***")
            print(awayteam)
            awayteam = awayteam.split(" ", 1)[1]
            scores = run_game(hometeam, awayteam)
            homescore = scores[0]
            awayscore = scores[1]
            
            if homescore > awayscore:
                schedulesheet.cell(row = j, column = i + 1).value = "W " + str(homescore) + "-" + str(awayscore)
            elif awayscore > homescore:
                schedulesheet.cell(row = j, column = i + 1).value = "L " + str(homescore) + "-" + str(awayscore)
            else:
                schedulesheet.cell(row = j, column = i + 1).value = "T " + str(homescore) + "-" + str(awayscore)

            #schedulesheet.cell(row = j, column = i+1).alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

            for n in range(2,66,2):
                if schedulesheet.cell(row = 1, column = n).value == awayteam:
                    if homescore > awayscore:
                        schedulesheet.cell(row = j, column = n+1).value = "L " + str(awayscore) + "-" + str(homescore)
                    elif awayscore > homescore:
                        schedulesheet.cell(row = j, column = n+1).value = "W " + str(awayscore) + "-" + str(homescore)
                    else:
                        schedulesheet.cell(row = j, column = n+1).value = "T " + str(awayscore) + "-" + str(homescore)
                #schedulesheet.cell(row = j, column = n+1).alignment = Alignment(horizontal = 'center', vertical = 'center')

            print(schedulesheet.cell(row = j, column = i+1).value)
            print("Game ID: " + scores[2])
            print("--------------------------------------------------------------")


#Read game results and add up wins, losses, points scored, and points allowed
for i in range(2,66,2):
    hometeam = schedulesheet.cell(row = 1, column = i).value
    for j in range(2,18):
        ihome = index_2d(teams, hometeam)[0]
        jhome = index_2d(teams, hometeam)[1]
        winsplit = schedulesheet.cell(row = j, column = i + 1).value.split(" ")
        if winsplit[0] == "W":
            teams[ihome][jhome].wins += 1
            teams[ihome][jhome].results[j-2] = "W"
        elif winsplit[0] == "L":
            teams[ihome][jhome].losses += 1
            teams[ihome][jhome].results[j-2] = "L"
        else:
            teams[ihome][jhome].ties += 1
            teams[ihome][jhome].results[j-2] = "T"
        scoresplit = winsplit[1].split("-")
        teams[ihome][jhome].pointsscored += int(scoresplit[0])
        teams[ihome][jhome].pointsallowed += int(scoresplit[1])

for i in range(8):
    for j in range(4):
        if (teams[i][j].ties > 0):
            teams[i][j].record = str(teams[i][j].wins) + "-" + str(teams[i][j].losses) + "-" + str(teams[i][j].ties)
        else:
            teams[i][j].record = str(teams[i][j].wins) + "-" + str(teams[i][j].losses)

diffsheet = wb["Year " + sys.argv[1] + " Point Differentials"]

diffteamlist = [Team() for x in range(32)]
for i in range(8):
    for j in range(4):
        diffteamlist[i*4+j] = teams[i][j]


#Fill out point differential ranking sheet
diffteamlist.sort(reverse = True, key = lambda Team: (Team.wins, Team.ties, (Team.pointsscored - Team.pointsallowed)))

for i in range(2,34):
    diffsheet.cell(row = i, column = 2).value = diffteamlist[i-2].name
    diffsheet.cell(row = i, column = 3).value = diffteamlist[i-2].pointsscored
    diffsheet.cell(row = i, column = 4).value = diffteamlist[i-2].pointsallowed
    diffsheet.cell(row = i, column = 5).value = diffteamlist[i-2].pointsscored - diffteamlist[i-2].pointsallowed
    diffsheet.cell(row = i, column = 6).value = diffteamlist[i-2].record


#Determine which teams are playoff-bound
for i in range(8):
    teams[i].sort(reverse = True, key = lambda Team: (Team.wins, Team.ties))

afcdivleaders = [Team() for i in range(4)]
nfcdivleaders = [Team() for i in range(4)]
afcwildcardcontenders = [Team() for i in range(8)]
nfcwildcardcontenders = [Team() for i in range(8)]

for i in range(8):
    playoffsort(teams[i])
    if i < 4:
        afcdivleaders[i] = teams[i][0]
        afcwildcardcontenders[i*2] = teams[i][1]
        afcwildcardcontenders[i*2+1] = teams[i][2]
    else:
        nfcdivleaders[i-4] = teams[i][0]
        nfcwildcardcontenders[(i-4)*2] = teams[i][1]
        nfcwildcardcontenders[(i-4)*2+1] = teams[i][2]

# print("")
# print("Sorting AFC Division Leaders")
# for i in range(4):
#     print(afcdivleaders[i].name + ": " + afcdivleaders[i].record)
playoffsort(afcdivleaders)
# print("")
# print("Sorting NFC Division Leaders")
# for i in range(4):
#     print(nfcdivleaders[i].name + ": " + nfcdivleaders[i].record)
playoffsort(nfcdivleaders)
# print("")
# print("Sorting AFC Wild Card Contenders")
# for i in range(8):
#     print(afcwildcardcontenders[i].name + ": " + afcwildcardcontenders[i].record)
playoffsort(afcwildcardcontenders)
# print("")
# print("Sorting NFC Wild Card Contenders")
# for i in range(8):
#     print(nfcwildcardcontenders[i].name + ": " + nfcwildcardcontenders[i].record)
playoffsort(nfcwildcardcontenders)

afcplayoffteams = [Team() for i in range(6)]
afcdivroundteams = [Team() for i in range(4)]
afcconfchampteams = [Team() for i in range(2)]
nfcplayoffteams = [Team() for i in range(6)]
nfcdivroundteams = [Team() for i in range(4)]
nfcconfchampteams = [Team() for i in range(2)]
superbowlteams = [Team() for i in range(2)]

for i in range(4):
    afcplayoffteams[i] = afcdivleaders[i]
    afcplayoffteams[i].playoffseed = i+1
    nfcplayoffteams[i] = nfcdivleaders[i]
    nfcplayoffteams[i].playoffseed = i+1

for i in range(4,6):
    afcplayoffteams[i] = afcwildcardcontenders[i-4]
    afcplayoffteams[i].playoffseed = i+1
    nfcplayoffteams[i] = nfcwildcardcontenders[i-4]
    nfcplayoffteams[i].playoffseed = i+1


#Fill out division standings sheet
divsheetname = "Year " + sys.argv[1] + " Division Standings"
standingssheet = wb[divsheetname]
teamctr = 0
for i in range(1,11,3):
    for j in range(2,6):
        standingssheet.cell(row = j, column = i).value = teams[teamctr][j-2].name
        standingssheet.cell(row = j, column = i+1).value = teams[teamctr][j-2].record
    for j in range(8,12):
        standingssheet.cell(row = j, column = i).value = teams[teamctr+4][j-8].name
        standingssheet.cell(row = j, column = i+1).value = teams[teamctr+4][j-8].record
    teamctr += 1;


#Fill out playoff bracket sheet
bracketname = "Year " + sys.argv[1] + " Playoffs"
bracket = wb[bracketname]
bracket.cell(5,1).value = "3"
bracket.cell(7,1).value = "6"
bracket.cell(5,2).value = afcplayoffteams[2].name
bracket.cell(7,2).value = afcplayoffteams[5].name
bracket.cell(11,1).value = "4"
bracket.cell(13,1).value = "5"
bracket.cell(11,2).value = afcplayoffteams[3].name
bracket.cell(13,2).value = afcplayoffteams[4].name

bracket.cell(5,11).value = "3"
bracket.cell(7,11).value = "6"
bracket.cell(5,10).value = nfcplayoffteams[2].name
bracket.cell(7,10).value = nfcplayoffteams[5].name
bracket.cell(11,11).value = "4"
bracket.cell(13,11).value = "5"
bracket.cell(11,10).value = nfcplayoffteams[3].name
bracket.cell(13,10).value = nfcplayoffteams[4].name

# hometeamrows = [5,11,5,11,2,16,2,16]
# playoffcols = [2,2,10,10,3,3,9,9]
# awayteamrows = [7,13,7,13,6,12,6,12]


#AFC Wild Card Round 3 Seed vs. 6 Seed Matchup
print("")
print("#-----Welcome to the NFL Playoffs!-----#")
print("")
print("#----------AFC Wild Card Round----------")
print("3 Seed: " + afcplayoffteams[2].name + " (" + afcplayoffteams[2].record + ")")
print("                     VS.                ")
print("6 Seed: " + afcplayoffteams[5].name + " (" + afcplayoffteams[5].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(afcplayoffteams[2].name, afcplayoffteams[5].name)
while scores[0] == scores[1]:
    scores = run_game(afcplayoffteams[2].name, afcplayoffteams[5].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + afcplayoffteams[2].name + " " + str(scores[0]) + "-" + str(scores[1]))
    afcdivroundteams[0] = afcplayoffteams[1]
    afcdivroundteams[1] = afcplayoffteams[2]
    afcdivroundteams[2] = afcplayoffteams[0]
    bracket.cell(6,3).value = afcplayoffteams[2].name
    bracket.cell(6,2).value = str(scores[0]) + "-" + str(scores[1])
    bracket.cell(2,3).value = afcplayoffteams[1].name
    bracket.cell(2,2).value = "2"
    bracket.cell(16,3).value = afcplayoffteams[0].name
    bracket.cell(16,2).value = "1"
else:
    print("Winner: " + afcplayoffteams[5].name + " " + str(scores[1]) + "-" + str(scores[0]))
    afcdivroundteams[0] = afcplayoffteams[0]
    afcdivroundteams[1] = afcplayoffteams[5]
    afcdivroundteams[2] = afcplayoffteams[1]
    bracket.cell(6,3).value = afcplayoffteams[5].name
    bracket.cell(6,2).value = str(scores[1]) + "-" + str(scores[0])
    bracket.cell(2,3).value = afcplayoffteams[0].name
    bracket.cell(2,2).value = "1"
    bracket.cell(16,3).value = afcplayoffteams[1].name
    bracket.cell(16,2).value = "2"
print("Game ID: " + scores[2])
print("----------------------------------------")


#AFC Wild Card Round 4 Seed vs. 5 Seed Matchup
print("")
print("4 Seed: " + afcplayoffteams[3].name + " (" + afcplayoffteams[3].record + ")")
print("                     VS.                ")
print("5 Seed: " + afcplayoffteams[4].name + " (" + afcplayoffteams[4].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(afcplayoffteams[3].name, afcplayoffteams[4].name)
while scores[0] == scores[1]:
    scores = run_game(afcplayoffteams[3].name, afcplayoffteams[4].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + afcplayoffteams[3].name + " " + str(scores[0]) + "-" + str(scores[1]))
    afcdivroundteams[3] = afcplayoffteams[3]
    bracket.cell(12,3).value = afcplayoffteams[3].name
    bracket.cell(12,2).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + afcplayoffteams[4].name + " " + str(scores[1]) + "-" + str(scores[0]))
    afcdivroundteams[3] = afcplayoffteams[4]
    bracket.cell(12,3).value = afcplayoffteams[4].name
    bracket.cell(12,2).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#NFC Wild Card Round 3 Seed vs. 6 Seed Matchup
print("")
print("#----------NFC Wild Card Round----------")
print("3 Seed: " + nfcplayoffteams[2].name + " (" + nfcplayoffteams[2].record + ")")
print("                     VS.                ")
print("6 Seed: " + nfcplayoffteams[5].name + " (" + nfcplayoffteams[5].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(nfcplayoffteams[2].name, nfcplayoffteams[5].name)
while scores[0] == scores[1]:
    scores = run_game(nfcplayoffteams[2].name, nfcplayoffteams[5].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + nfcplayoffteams[2].name + " " + str(scores[0]) + "-" + str(scores[1]))
    nfcdivroundteams[0] = nfcplayoffteams[1]
    nfcdivroundteams[1] = nfcplayoffteams[2]
    nfcdivroundteams[2] = nfcplayoffteams[0]
    bracket.cell(6,9).value = nfcplayoffteams[2].name
    bracket.cell(6,10).value = str(scores[0]) + "-" + str(scores[1])
    bracket.cell(2,9).value = nfcplayoffteams[1].name
    bracket.cell(2,10).value = "2"
    bracket.cell(16,9).value = nfcplayoffteams[0].name
    bracket.cell(16,10).value = "1"
else:
    print("Winner: " + nfcplayoffteams[5].name + " " + str(scores[1]) + "-" + str(scores[0]))
    nfcdivroundteams[0] = nfcplayoffteams[0]
    nfcdivroundteams[1] = nfcplayoffteams[5]
    nfcdivroundteams[2] = nfcplayoffteams[1]
    bracket.cell(6,9).value = nfcplayoffteams[5].name
    bracket.cell(6,10).value = str(scores[1]) + "-" + str(scores[0])
    bracket.cell(2,9).value = nfcplayoffteams[0].name
    bracket.cell(2,10).value = "1"
    bracket.cell(16,9).value = nfcplayoffteams[1].name
    bracket.cell(16,10).value = "2"
print("Game ID: " + scores[2])
print("----------------------------------------")


#NFC Wild Card Round 4 Seed vs. 5 Seed Matchup
print("")
print("4 Seed: " + nfcplayoffteams[3].name + " (" + nfcplayoffteams[3].record + ")")
print("                    VS.                ")
print("5 Seed: " + nfcplayoffteams[4].name + " (" + nfcplayoffteams[4].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(nfcplayoffteams[3].name, nfcplayoffteams[4].name)
while scores[0] == scores[1]:
    scores = run_game(nfcplayoffteams[3].name, nfcplayoffteams[4].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + nfcplayoffteams[3].name + " " + str(scores[0]) + "-" + str(scores[1]))
    nfcdivroundteams[3] = nfcplayoffteams[3]
    bracket.cell(12,9).value = nfcplayoffteams[3].name
    bracket.cell(12,10).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + nfcplayoffteams[4].name + " " + str(scores[1]) + "-" + str(scores[0]))
    nfcdivroundteams[3] = nfcplayoffteams[4]
    bracket.cell(12,9).value = nfcplayoffteams[4].name
    bracket.cell(12,10).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#AFC Divisional Round Higher Matchup
print("")
print("#----------AFC Divisional Round----------")
print(str(afcdivroundteams[0].playoffseed) + " Seed: " + afcdivroundteams[0].name + " (" + afcdivroundteams[0].record + ")")
print("                     VS.                ")
print(str(afcdivroundteams[1].playoffseed) + " Seed: " + afcdivroundteams[1].name + " (" + afcdivroundteams[1].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(afcdivroundteams[0].name, afcdivroundteams[1].name)
while scores[0] == scores[1]:
    scores = run_game(afcdivroundteams[0].name, afcdivroundteams[1].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + afcdivroundteams[0].name + " " + str(scores[0]) + "-" + str(scores[1]))
    afcconfchampteams[0] = afcdivroundteams[0]
    bracket.cell(4,4).value = afcdivroundteams[0].name
    bracket.cell(4,3).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + afcdivroundteams[1].name + " " + str(scores[1]) + "-" + str(scores[0]))
    afcconfchampteams[0] = afcdivroundteams[1]
    bracket.cell(4,4).value = afcdivroundteams[1].name
    bracket.cell(4,3).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#AFC Divisional Round Lower Matchup
print("")
print(str(afcdivroundteams[2].playoffseed) + " Seed: " + afcdivroundteams[2].name + " (" + afcdivroundteams[2].record + ")")
print("                     VS.                ")
print(str(afcdivroundteams[3].playoffseed) + " Seed: " + afcdivroundteams[3].name + " (" + afcdivroundteams[3].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(afcdivroundteams[2].name, afcdivroundteams[3].name)
while scores[0] == scores[1]:
    scores = run_game(afcdivroundteams[2].name, afcdivroundteams[3].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + afcdivroundteams[2].name + " " + str(scores[0]) + "-" + str(scores[1]))
    afcconfchampteams[1] = afcdivroundteams[2]
    bracket.cell(14,4).value = afcdivroundteams[2].name
    bracket.cell(14,3).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + afcdivroundteams[3].name + " " + str(scores[1]) + "-" + str(scores[0]))
    afcconfchampteams[1] = afcdivroundteams[3]
    bracket.cell(14,4).value = afcdivroundteams[3].name
    bracket.cell(14,3).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#NFC Divisional Round Higher Matchup
print("")
print("#----------NFC Divisional Round----------")
print(str(nfcdivroundteams[0].playoffseed) + " Seed: " + nfcdivroundteams[0].name + " (" + nfcdivroundteams[0].record + ")")
print("                     VS.                ")
print(str(nfcdivroundteams[1].playoffseed) + " Seed: " + nfcdivroundteams[1].name + " (" + nfcdivroundteams[1].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(nfcdivroundteams[0].name, nfcdivroundteams[1].name)
while scores[0] == scores[1]:
    scores = run_game(nfcdivroundteams[0].name, nfcdivroundteams[1].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + nfcdivroundteams[0].name + " " + str(scores[0]) + "-" + str(scores[1]))
    nfcconfchampteams[0] = nfcdivroundteams[0]
    bracket.cell(4,8).value = nfcdivroundteams[0].name
    bracket.cell(4,9).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + nfcdivroundteams[1].name + " " + str(scores[1]) + "-" + str(scores[0]))
    nfcconfchampteams[0] = nfcdivroundteams[1]
    bracket.cell(4,8).value = nfcdivroundteams[1].name
    bracket.cell(4,9).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#NFC Divisional Round Lower Matchup
print("")
print(str(nfcdivroundteams[2].playoffseed) + " Seed: " + nfcdivroundteams[2].name + " (" + nfcdivroundteams[2].record + ")")
print("                     VS.                ")
print(str(nfcdivroundteams[3].playoffseed) + " Seed: " + nfcdivroundteams[3].name + " (" + nfcdivroundteams[3].record + ")")
input("Press Enter to run game:")
print("")

scores = run_game(nfcdivroundteams[2].name, nfcdivroundteams[3].name)
while scores[0] == scores[1]:
    scores = run_game(nfcdivroundteams[2].name, nfcdivroundteams[3].name) #Can't have a tie, so does it again if there is one
if scores[0] > scores[1]:
    print("Winner: " + nfcdivroundteams[2].name + " " + str(scores[0]) + "-" + str(scores[1]))
    nfcconfchampteams[1] = nfcdivroundteams[2]
    bracket.cell(14,8).value = nfcdivroundteams[2].name
    bracket.cell(14,9).value = str(scores[0]) + "-" + str(scores[1])
else:
    print("Winner: " + nfcdivroundteams[3].name + " " + str(scores[1]) + "-" + str(scores[0]))
    nfcconfchampteams[1] = nfcdivroundteams[3]
    bracket.cell(14,8).value = nfcdivroundteams[3].name
    bracket.cell(14,9).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#AFC Conference Championship
print("")
print("#----------AFC CONFERENCE CHAMPIONSHIP----------#")
print(str(afcconfchampteams[0].playoffseed) + " Seed: " + afcconfchampteams[0].name + " (" + afcconfchampteams[0].record + ")")
print("                     VS.                 ")
print(str(afcconfchampteams[1].playoffseed) + " Seed: " + afcconfchampteams[1].name + " (" + afcconfchampteams[1].record + ")")
input("Press Enter to run game:")
print("")

if (afcconfchampteams[0].playoffseed > afcconfchampteams[1].playoffseed):
    scores = run_game(afcconfchampteams[0].name, afcconfchampteams[1].name)
    firstTeamHome = True 
else:
    scores = run_game(afcconfchampteams[0].name, afcconfchampteams[1].name)
    firstTeamHome = False
while scores[0] == scores[1]:
    if (firstTeamHome == True):
        scores = run_game(afcconfchampteams[0].name, afcconfchampteams[1].name)
    else:
        scores = run_game(afcconfchampteams[1].name, afcconfchampteams[0].name) #Can't have a tie, so does it again if there is one
if firstTeamHome == True:
    if scores[0] > scores[1]:
        print("Winner: " + afcconfchampteams[0].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlteams[0] = afcconfchampteams[0]
        bracket.cell(9,5).value = afcconfchampteams[0].name
        bracket.cell(9,4).value = str(scores[0]) + "-" + str(scores[1])
    else:
        print("Winner: " + afcconfchampteams[1].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlteams[0] = afcconfchampteams[1]
        bracket.cell(9,5).value = afcconfchampteams[1].name
        bracket.cell(9,4).value = str(scores[1]) + "-" + str(scores[0])
else:
    if scores[0] > scores[1]:
        print("Winner: " + afcconfchampteams[1].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlteams[0] = afcconfchampteams[1]
        bracket.cell(9,5).value = afcconfchampteams[1].name
        bracket.cell(9,4).value = str(scores[0]) + "-" + str(scores[1])
    else:
        print("Winner: " + afcconfchampteams[0].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlteams[0] = afcconfchampteams[0]
        bracket.cell(9,5).value = afcconfchampteams[0].name
        bracket.cell(9,4).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#NFC Conference Championship
print("")
print("#----------NFC CONFERENCE CHAMPIONSHIP----------#")
print(str(nfcconfchampteams[0].playoffseed) + " Seed: " + nfcconfchampteams[0].name + " (" + nfcconfchampteams[0].record + ")")
print("                     VS.                 ")
print(str(nfcconfchampteams[1].playoffseed) + " Seed: " + nfcconfchampteams[1].name + " (" + nfcconfchampteams[1].record + ")")
input("Press Enter to run game:")
print("")

if (nfcconfchampteams[0].playoffseed > nfcconfchampteams[1].playoffseed):
    scores = run_game(nfcconfchampteams[0].name, nfcconfchampteams[1].name)
    firstTeamHome = True 
else:
    scores = run_game(nfcconfchampteams[0].name, nfcconfchampteams[1].name)
    firstTeamHome = False
while scores[0] == scores[1]:
    if (firstTeamHome == True):
        scores = run_game(nfcconfchampteams[0].name, nfcconfchampteams[1].name)
    else:
        scores = run_game(nfcconfchampteams[1].name, nfcconfchampteams[0].name) #Can't have a tie, so does it again if there is one
if firstTeamHome == True:
    if scores[0] > scores[1]:
        print("Winner: " + nfcconfchampteams[0].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlteams[1] = nfcconfchampteams[0]
        bracket.cell(9,7).value = nfcconfchampteams[0].name
        bracket.cell(9,8).value = str(scores[0]) + "-" + str(scores[1])
    else:
        print("Winner: " + nfcconfchampteams[1].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlteams[1] = nfcconfchampteams[1]
        bracket.cell(9,7).value = nfcconfchampteams[1].name
        bracket.cell(9,8).value = str(scores[1]) + "-" + str(scores[0])
else:
    if scores[0] > scores[1]:
        print("Winner: " + nfcconfchampteams[1].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlteams[1] = nfcconfchampteams[1]
        bracket.cell(9,7).value = nfcconfchampteams[1].name
        bracket.cell(9,8).value = str(scores[0]) + "-" + str(scores[1])
    else:
        print("Winner: " + nfcconfchampteams[0].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlteams[1] = nfcconfchampteams[0]
        bracket.cell(9,7).value = nfcconfchampteams[0].name
        bracket.cell(9,8).value = str(scores[1]) + "-" + str(scores[0])
print("Game ID: " + scores[2])
print("----------------------------------------")


#Supa Bowl Babyyyyy
superbowlnumerals = ['I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII','XIII','XIV','XV','XVI','XVII','XVIII','XIX','XX','XXI','XXII','XXIII','XXIV','XXV','XXVI','XXVII','XXVIII','XXIX','XXX','XXXI','XXXII','XXXIII','XXXIV','XXXV','XXXVI','XXXVII','XXXVIII','XXXIX','XL','XLI','XLII','XLIII','XLIV','XLV','XLVI','XLVII','XLVIII','XLIX','50','LI','LII','LIII','LIV']
print("")
print("#---------SUPER BOWL " + superbowlnumerals[int(sys.argv[1])-1] + "-----------#")
print(str(superbowlteams[0].playoffseed) + " Seed: " + superbowlteams[0].name + " (" + superbowlteams[0].record + ")")
print("                     VS.                 ")
print(str(superbowlteams[1].playoffseed) + " Seed: " + superbowlteams[1].name + " (" + superbowlteams[1].record + ")")
input("Press Enter to run game:")
print("")

superbowlsheet = wb["Super Bowl Results"]
if (superbowlteams[0].playoffseed > superbowlteams[1].playoffseed):
    scores = run_game(superbowlteams[0].name, superbowlteams[1].name)
    firstTeamHome = True 
else:
    scores = run_game(superbowlteams[0].name, superbowlteams[1].name)
    firstTeamHome = False
while scores[0] == scores[1]:
    if (firstTeamHome == True):
        scores = run_game(superbowlteams[0].name, superbowlteams[1].name)
    else:
        scores = run_game(superbowlteams[1].name, superbowlteams[0].name) #Can't have a tie, so does it again if there is one
superbowlsheet.cell(row = int(sys.argv[1])+1, column = 1).value = superbowlnumerals[int(sys.argv[1])-1]
if firstTeamHome == True:
    if scores[0] > scores[1]:
        print("Winner: " + superbowlteams[0].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlwinner = superbowlteams[0]
        bracket.cell(18,5).value = superbowlteams[0].name
        bracket.cell(10,6).value = str(scores[0]) + "-" + str(scores[1])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 2).value = superbowlteams[0].name
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 3).value = str(scores[0]) + "-" + str(scores[1])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 4).value = superbowlteams[1].name
    else:
        print("Winner: " + superbowlteams[1].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlwinner = superbowlteams[1]
        bracket.cell(18,5).value = superbowlteams[1].name
        bracket.cell(10,6).value = str(scores[1]) + "-" + str(scores[0])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 2).value = superbowlteams[1].name
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 3).value = str(scores[1]) + "-" + str(scores[0])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 4).value = superbowlteams[0].name
else:
    if scores[0] > scores[1]:
        print("Winner: " + superbowlteams[1].name + " " + str(scores[0]) + "-" + str(scores[1]))
        superbowlwinner = superbowlteams[1]
        bracket.cell(18,5).value = superbowlteams[1].name
        bracket.cell(10,6).value = str(scores[0]) + "-" + str(scores[1])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 2).value = superbowlteams[1].name
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 3).value = str(scores[0]) + "-" + str(scores[1])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 4).value = superbowlteams[0].name
    else:
        print("Winner: " + superbowlteams[0].name + " " + str(scores[1]) + "-" + str(scores[0]))
        superbowlwinner = superbowlteams[0]
        bracket.cell(18,5).value = superbowlteams[0].name
        bracket.cell(10,6).value = str(scores[1]) + "-" + str(scores[0])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 2).value = superbowlteams[0].name
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 3).value = str(scores[1]) + "-" + str(scores[0])
        superbowlsheet.cell(row = int(sys.argv[1])+1, column = 4).value = superbowlteams[1].name
print("Game ID: " + scores[2])
superbowlsheet.cell(row = int(sys.argv[1]) + 1, column = 5).value = scores[2]
print("----------------------------------------")


#Fill out All-Time Win-Loss Record sheet
alltimesheet = wb["All-Time Results"]

for i in range(32):
    for j in range(2, 34):
        if alltimesheet.cell(row = j, column = 2).value == diffteamlist[i].name:
            alltimesheet.cell(row = j, column = 4).value = str(int(alltimesheet.cell(row = j, column = 4).value) + diffteamlist[i].wins)
            alltimesheet.cell(row = j, column = 5).value = str(int(alltimesheet.cell(row = j, column = 5).value) + diffteamlist[i].losses)
            alltimesheet.cell(row = j, column = 6).value = str(int(alltimesheet.cell(row = j, column = 6).value) + diffteamlist[i].ties)
            for k in range(6):
                if diffteamlist[i].name == afcplayoffteams[k].name:
                    alltimesheet.cell(row = j, column = 8).value = str(int(alltimesheet.cell(row = j, column = 8).value) + 1)
                elif diffteamlist[i].name == nfcplayoffteams[k].name:
                    alltimesheet.cell(row = j, column = 8).value = str(int(alltimesheet.cell(row = j, column = 8).value) + 1)
            for k in range(2):
                if diffteamlist[i].name == afcconfchampteams[k].name:
                    alltimesheet.cell(row = j, column = 9).value = str(int(alltimesheet.cell(row = j, column = 9).value) + 1)
                elif diffteamlist[i].name == nfcconfchampteams[k].name:
                    alltimesheet.cell(row = j, column = 9).value = str(int(alltimesheet.cell(row = j, column = 9).value) + 1)
                if diffteamlist[i].name == superbowlteams[k].name:
                    alltimesheet.cell(row = j, column = 10).value = str(int(alltimesheet.cell(row = j, column = 10).value) + 1)
            if diffteamlist[i].name == superbowlwinner.name:
                alltimesheet.cell(row = j, column = 11).value = str(int(alltimesheet.cell(row = j, column = 11).value) + 1)


#Add top-ranked regular season team to Yearly Top-Ranked Teams sheet
toprankedsheet = wb["Yearly Top-Ranked Teams"]
toprankedsheet.cell(row = int(sys.argv[1])+1, column = 1).value = "Year " + sys.argv[1]
toprankedsheet.cell(row = int(sys.argv[1])+1, column = 2).value = diffteamlist[0].name
toprankedsheet.cell(row = int(sys.argv[1])+1, column = 3).value = diffteamlist[0].record
toprankedsheet.cell(row = int(sys.argv[1])+1, column = 4).value = diffteamlist[0].pointsscored - diffteamlist[0].pointsallowed


#wb.remove_sheet(schedulesheetname)
wb.save(filepath)
